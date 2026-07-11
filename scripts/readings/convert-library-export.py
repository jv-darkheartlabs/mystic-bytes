#!/usr/bin/env python3
"""Convert mystic-bytes Goodreads CSV export to .xlsx and tab-delimited .txt (clean UTF-8)."""
import csv
import re
import sys
import unicodedata
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

DEFAULT_CSV = Path.home() / "Downloads/mystic-bytes_library_export.csv"
DEFAULT_XLSX = Path.home() / "Downloads/mystic-bytes_library_export.xlsx"
DEFAULT_TXT = Path.home() / "Downloads/mystic-bytes_library_export.txt"

REVIEW_COL = 13
WIDE_COLS = {0: 42, 1: 28, 5: 24, 12: 36, REVIEW_COL: 80}

SMART_QUOTES = str.maketrans(
    {
        "\u2018": "'",
        "\u2019": "'",
        "\u201a": "'",
        "\u201b": "'",
        "\u201c": '"',
        "\u201d": '"',
        "\u201e": '"',
        "\u201f": '"',
        "\u2013": "-",
        "\u2014": "-",
        "\u2026": "...",
        "\u2022": "-",
    }
)


def sanitize_utf8(value: object) -> str:
    if value is None:
        return ""
    s = str(value).replace("\ufeff", "")
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", s)
    s = s.translate(SMART_QUOTES)
    s = re.sub(r"[\u200b-\u200d\u2060]", "", s)
    s = unicodedata.normalize("NFC", s)
    return s.encode("utf-8", "strict").decode("utf-8")


def read_csv(path: Path) -> list[list[str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        rows = [[sanitize_utf8(cell) for cell in row] for row in csv.reader(f)]
    if not rows:
        raise SystemExit(f"Empty CSV: {path}")
    return rows


def write_txt(rows: list[list[str]], path: Path) -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(
            f,
            delimiter="\t",
            quoting=csv.QUOTE_MINIMAL,
            lineterminator="\n",
        )
        writer.writerows(rows)


def write_xlsx(rows: list[list[str]], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Library Export"

    header_font = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    top = Alignment(vertical="top")

    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = header_font
                cell.alignment = Alignment(wrap_text=True, vertical="center")
            elif c_idx - 1 == REVIEW_COL:
                cell.alignment = wrap
            else:
                cell.alignment = top

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx, width in WIDE_COLS.items():
        ws.column_dimensions[get_column_letter(col_idx + 1)].width = width

    for c in range(1, len(rows[0]) + 1):
        letter = get_column_letter(c)
        if letter not in ws.column_dimensions or ws.column_dimensions[letter].width is None:
            ws.column_dimensions[letter].width = 14

    wb.save(path)


def validate_utf8_text(path: Path) -> None:
    raw = path.read_bytes()
    raw.decode("utf-8", "strict")
    if raw.startswith(b"\xef\xbb\xbf"):
        raise ValueError(f"{path.name} has UTF-8 BOM (should be BOM-free)")


def main() -> None:
    csv_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_CSV
    xlsx_path = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_XLSX
    txt_path = Path(sys.argv[3]) if len(sys.argv) > 3 else DEFAULT_TXT

    rows = read_csv(csv_path)
    print(f"Read {len(rows) - 1} data rows from {csv_path}")

    # Rewrite CSV clean (no BOM, sanitized, LF)
    with csv_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, lineterminator="\n")
        writer.writerows(rows)
    validate_utf8_text(csv_path)
    print(f"Rewrote clean CSV → {csv_path}")

    write_xlsx(rows, xlsx_path)
    print(f"Wrote {xlsx_path}")

    write_txt(rows, txt_path)
    validate_utf8_text(txt_path)
    print(f"Wrote {txt_path} (UTF-8, no BOM, LF)")


if __name__ == "__main__":
    main()
